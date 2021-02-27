import openpyxl


def write_xlsx(data: list[dict], filename: str) -> bool:
    wb = openpyxl.Workbook()
    ws = wb.active
    
    print(f'Len of data is {len(data)}')

    column_names = list()
    for el in data:
        for key in el:
            if key not in column_names:
                column_names.append(key)
                
    for j, el in enumerate(column_names):
        ws[F'{chr(ord("A") + j)}{1}'] = el
    
    for i, row in enumerate(data, 1):
        for j, el in enumerate(column_names):
            try:
                ws[f'{chr(ord("A") + j)}{i + 1}'] = data[i - 1][el]
            except KeyError:
                pass
    
    wb.save(filename)
    
    return True